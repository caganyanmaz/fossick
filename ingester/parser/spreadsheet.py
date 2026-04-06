from __future__ import annotations

from pathlib import Path
from typing import Any

from loguru import logger

from ingester.parser.base import BaseParser, ParsedDocument

_EXTENSIONS = {"xlsx", "xls", "csv"}


def _rows_to_markdown(headers: list[str], rows: list[list[Any]]) -> str:
    """Convert tabular data to a markdown table string."""
    if not headers:
        return ""
    col_widths = [max(len(str(h)), 3) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            if i < len(col_widths):
                col_widths[i] = max(col_widths[i], len(str(cell) if cell is not None else ""))

    def fmt_row(cells: list[Any]) -> str:
        padded = []
        for i, c in enumerate(cells):
            s = str(c) if c is not None else ""
            w = col_widths[i] if i < len(col_widths) else len(s)
            padded.append(s.ljust(w))
        return "| " + " | ".join(padded) + " |"

    separator = "| " + " | ".join("-" * w for w in col_widths) + " |"
    lines = [fmt_row(headers), separator] + [fmt_row(row) for row in rows]
    return "\n".join(lines)


class SpreadsheetParser(BaseParser):
    def can_parse(self, path: str) -> bool:
        return Path(path).suffix.lstrip(".").lower() in _EXTENSIONS

    def parse(self, path: str) -> ParsedDocument:
        p = Path(path)
        filetype = p.suffix.lstrip(".").lower()
        stat = p.stat()

        if filetype == "csv":
            return self._parse_csv(p, stat)
        return self._parse_excel(p, filetype, stat)

    def _parse_csv(self, p: Path, stat: Any) -> ParsedDocument:
        import pandas as pd

        try:
            df = pd.read_csv(str(p), nrows=1000)
            text = df.to_markdown(index=False) or ""
            meta: dict[str, Any] = {
                "filename": p.name,
                "path": str(p),
                "filetype": "csv",
                "size_bytes": stat.st_size,
                "mtime": stat.st_mtime,
                "row_count": len(df),
                "column_names": list(df.columns),
            }
        except Exception as exc:
            logger.warning("CSV parse failed for {}: {}", p, exc)
            text = p.read_text(errors="replace")
            meta = {
                "filename": p.name,
                "path": str(p),
                "filetype": "csv",
                "size_bytes": stat.st_size,
                "mtime": stat.st_mtime,
                "row_count": 0,
                "column_names": [],
            }
        return ParsedDocument(text=text, metadata=meta, source_path=str(p))

    def _parse_excel(self, p: Path, filetype: str, stat: Any) -> ParsedDocument:
        import openpyxl

        try:
            wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            sections: list[str] = []

            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                all_rows = []
                for i, row in enumerate(rows_iter):
                    if i >= 1001:
                        break
                    all_rows.append(list(row))

                if not all_rows:
                    continue

                headers = [str(c) if c is not None else "" for c in all_rows[0]]
                data_rows = [list(r) for r in all_rows[1:1001]]
                table = _rows_to_markdown(headers, data_rows)
                sections.append(f"## Sheet: {sheet_name}\n\n{table}")

            wb.close()
            text = "\n\n".join(sections)
            meta: dict[str, Any] = {
                "filename": p.name,
                "path": str(p),
                "filetype": filetype,
                "size_bytes": stat.st_size,
                "mtime": stat.st_mtime,
                "sheet_names": sheet_names,
            }
        except Exception as exc:
            logger.warning("Excel parse failed for {}: {}", p, exc)
            text = ""
            meta = {
                "filename": p.name,
                "path": str(p),
                "filetype": filetype,
                "size_bytes": stat.st_size,
                "mtime": stat.st_mtime,
                "sheet_names": [],
            }
        return ParsedDocument(text=text, metadata=meta, source_path=str(p))
